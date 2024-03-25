import shutil
import datetime
import os
from tkinter.filedialog import askdirectory
import openpyxl


def remove_values_from_list(the_list, val):
    return [value for value in the_list if value != val]


def remove_comments(cals, num_lines):
    for index in range(0, num_lines):
        tmp_str = cals[index]
        str_idx = tmp_str.find('%')
        if str_idx != -1:
            tmp_str = tmp_str[0:str_idx - 1] + '\n'
            cals[index] = tmp_str
    return cals


def setup_out_file():
    # Open Files in w+ mode-> reading and writing, existing content is cleared, a new empty
    # file is created if it doesn't exist, file pointer is positioned at the beginning
    with open(tempFile, 'w+') as file:
        pass

    with open(calibrationTxtFile, 'w+') as file:
        pass

    with open(outFile, 'w+') as f5:
        datestamp = datetime.datetime.now()
        date = datestamp.strftime("%m/%d/%Y")
        time = datestamp.strftime('%H:%M:%S\n')
        f5.write(date + time)
    return date, time


def multi_line_cal_to_single_line(calibrations, num_lines):
    for index in range(num_lines, 0, -1):
        cur_str = calibrations[index]
        closed_bracket = cur_str.find(']')
        open_bracket = cur_str.find('[')
        count = 0
        # if closedBracket != -1 and openBracket == -1: # Detecting multiple lines (note reading from the bottom up)
        if closed_bracket != -1 and open_bracket == -1:  # Detecting multiple lines (note reading from the bottom up)
            while open_bracket == -1:
                # move current index-count to index-count-1
                calibrations[index - count - 1] = calibrations[index - count - 1].strip() + ' ; ' + calibrations[
                    index - count]
                # remove data from index-count
                calibrations[index - count] = 'DELETE'

                # conditions for the next check
                count = count + 1
                cur_str2 = calibrations[index - count]
                open_bracket = cur_str2.find('[')
            index = index - count
    calibrations = remove_values_from_list(calibrations, 'DELETE')


def write_calibrations_file(calibrations):
    num_lines = sum(1 for line in calibrations) - 1
    with open(calibrationTxtFile, 'w') as f2:
        for index in range(0, num_lines):
            f2.write(calibrations[index])


def search_cal_file_for_referenced_cal_names():

    cal_array_per_file = []
    cal_names = []
    array = []

    with open(outFile, 'a') as f0:
        with open(referenceFile, 'r') as f2:
            f0.write('Data_based_on_input_file: ' + inputFile + '\n')
            for index in f2:
                cal_full_names = index.splitlines()
                for index2 in range(0, sum(1 for line in cals_text) - 1):
                    str2 = (cals_text[index2]).lower()
                    cal_text_indexes = str2.find((cal_full_names[0].strip()).lower())

                    # split header from string, header is separated by '.'
                    if '.' not in str2:
                        continue
                    else:
                        header, str2 = str2.split('.', 1)
                    split_str2 = str2.split()

                    if cal_text_indexes != -1:
                        # Make array containing all cleaned up cal names
                        if split_str2[0] not in cal_names:
                            cal_names.append(split_str2[0])
                        # Write to out.txt
                        f0.write(str2)
                        # Cal names and values per file
                        if str2 not in cal_array_per_file:
                            cal_array_per_file.append(str2)
                        flag = 1
                        break
                    else:
                        flag = 0
                if flag < 1:  # not found
                    cal_not_found_message_for_file = str(cal_full_names[0] + '=\tNOT_FOUND\n')
                    cal_not_found_message_for_list = str(cal_full_names[0] + '=NOT_FOUND')
                    if cal_not_found_message_for_list not in cal_array_per_file:
                        cal_array_per_file.append(cal_not_found_message_for_list)
                    f0.write(cal_not_found_message_for_file)
                    array.append(cal_full_names[0])
                else:
                    flag = 0  # found string in nested loop
            f0.write('NEXT_FILE\n')
    return cal_names, cal_array_per_file


def size_columns_to_fit(worksheet):
    for ind, column in enumerate(worksheet.columns):
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        if ind == 0:
            adjusted_width = max_length  # * 1.2
        else:
            adjusted_width = 14
        worksheet.column_dimensions[column_letter].width = adjusted_width


def write_to_excel(array_of_all_file_cal_arrays, date, time):

    # Initializations vars needed for this function
    workbook = openpyxl.Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.title = 'Scalar Cals'
    array_indexes_all_files = []
    cal_arrays_inc_not_found = []
    array_of_scalar_cals = []
    array_of_array_cals = []
    array_of_arrays_of_cals = []
    scalar_cal_arrays_values_only = []
    array_values_organized_by_name = []
    array_cal_names = []
    loop_counter = 0

    # Populate top of sheet with date and time
    worksheet.cell(1, 1, date)
    worksheet.cell(1, 2, time)

    # Loop through list of cal lists to find array cals, make list of indexes of array cals for each file
    for i1, cal_array_per_m_file in enumerate(array_of_all_file_cal_arrays):
        array_indexes_per_file = []
        for i2, calibration in enumerate(cal_array_per_m_file):
            if '[' and ']' in calibration:
                array_indexes_per_file.append(i2)
        array_indexes_all_files.append(array_indexes_per_file)

    # Find longest list of array cals (most will be the same as longest, this is to get
    # index values so calibrations that are NOT_FOUND at those same indexes can be included
    max_list = max(array_indexes_all_files, key=len)
    # Sort list in descending order so when we remove from those indexes, numbers don't shift
    max_list.sort(reverse=True)

    # Remove cals corresponding to cal arrays, using indexes collected above
    # Assemble list containing only scalar cals, assemble list containing only array cals
    scalar_cal_names = cal_names_only
    for cal_array_per_m_file in array_of_all_file_cal_arrays:
        temp_cal_array = cal_array_per_m_file
        # Enumerate backwards so index counts down and removes from high index first
        for i2, cal in reversed(list(enumerate(cal_array_per_m_file))):
            for i3, value in enumerate(max_list):
                if int(i2) == int(value):
                    cal_arrays_inc_not_found.append(cal)
                    temp_cal_array.pop(i2)
                    loop_counter += 1
                    # Lots of m files to go through but only need to remove indexes from cal names once
                    if loop_counter <= len(max_list):
                        array_cal_names.append(scalar_cal_names[i2])
                        scalar_cal_names.pop(i2)
        array_of_scalar_cals.append(temp_cal_array)
        array_of_array_cals.append(cal_arrays_inc_not_found)
        cal_arrays_inc_not_found = []
    array_cal_names.reverse()

    # Split values for scalar cals off into their own list
    for scalar_cal_array in array_of_scalar_cals:
        scalar_temp_array_per_file = []
        for scalar_cal in scalar_cal_array:
            scalar_cal = scalar_cal.replace('\n', '')
            scalar_cal = scalar_cal.replace(';', '')
            split_scalar = scalar_cal.split('=')
            for item in split_scalar:
                item = item.replace('=', '')
                item.strip()
            scalar_temp_array_per_file.append(split_scalar[-1])
        scalar_cal_arrays_values_only.append(scalar_temp_array_per_file)

    # Split Values for array cals off into their own list
    for array_cal_array in array_of_array_cals:
        for array_cal in array_cal_array:
            array_cal = array_cal.replace('\n', '')
            array_cal = array_cal.replace(';', '')
            split_array_cal = array_cal.split('=')
            for thing in split_array_cal:
                thing = thing.replace('=', '')
                thing.strip()
            array_of_arrays_of_cals.append(split_array_cal[-1])

    # Take every nth element from list of array values and sort by cal name into new lists
    for i4, cal_name in enumerate(array_cal_names):
        array_values_organized_by_name.append(array_of_arrays_of_cals[i4::len(array_cal_names)])

    # Populate cal names into first column of first sheet
    for i5, cal_name in enumerate(scalar_cal_names):
        worksheet.cell(5+i5, 1, cal_name)

    # Populate File names and scalar cal values into first sheet, if there are scalar values
    if len(scalar_cal_names) == 0:
        worksheet.cell(5,1, 'NO SCALAR CALIBRATIONS FOUND')
        worksheet.cell(6,1,'ARRAY CALS FOUND IN FOLLOWING SHEETS')
    else:
        for i0, file in enumerate(files):
            worksheet.cell(3, 2 + i0, str(file))
        for i6, array_of_scalar_cal_values in enumerate(scalar_cal_arrays_values_only):
            for i7, scalar_cal_value in enumerate(array_of_scalar_cal_values):
                try:
                    scalar_cal_value = round(float(scalar_cal_value), 4)
                except:
                    pass
                worksheet.cell(5+i7, 2+i6, scalar_cal_value)
    size_columns_to_fit(worksheet)

    # Create a new sheet for each array cal name
    for i8, array_cal_name in enumerate(array_cal_names):
        # Excel sheets can only have 31char length titles, cut down if too long
        if len(array_cal_name) > 31:
            workbook.create_sheet(title=array_cal_name[:30], index=(i8 + 1))
        else:
            workbook.create_sheet(title=array_cal_name, index=(i8 + 1))
        worksheet_names = workbook.sheetnames
        worksheet = workbook[worksheet_names[i8 + 1]]
        # Put cal name at top of sheet
        worksheet.cell(1, 1, array_cal_name)
        worksheet.cell(2,1,'INDEX:')
        # Put file names in first column of sheet and resize columns to fit
        for i9, file in enumerate(files):
            worksheet.cell(4+i9, 1, file)
        size_columns_to_fit(worksheet)

    # Populate sheets with array data
    for i10, array_of_arrays in enumerate(array_values_organized_by_name):
        worksheet = workbook[worksheet_names[i10 + 1]]
        for i11, array in enumerate(array_of_arrays):
            array = array.replace('[', '')
            array = array.replace(']', '')
            array = array.strip()
            cal_array_split_into_values = array.split()
            for i12, array_cal_value in enumerate(cal_array_split_into_values):
                if i11 == 0:
                    worksheet.cell(2, 2+i12, int(i12))
                try:
                    array_cal_value = round(float(array_cal_value), 4)
                except:
                    pass
                worksheet.cell(4+i11, 2+i12, array_cal_value)
    workbook.save('Calibrations.xlsx')


# Author: Cody Palmer
# Last modified by: Cameron Floyd
# Last modified Date: 3-22-2024


root_dir = askdirectory(title='Select Folder that Contains your .m Files')
wrong_file_counter = 0

### File Preparation ###

### Initialize Files Needed ###

tempFile = 'temp.txt'
calibrationTxtFile = 'calibrations.txt'
outFile = 'out.txt'
referenceFile = 'references.txt'

arrays_of_all_file_cal_arrays = []

# create/overwrite output file and add timedate to the top
date_, time_ = setup_out_file()

# loop to iterate through files in selected folder
for subdir, dirs, files in os.walk(root_dir):
    for inputFile in files:
        if str(inputFile).endswith('.py'):
            continue
        elif str(inputFile).endswith('.txt'):
            continue
        elif str(inputFile).endswith('.m') is not True:
            wrong_file_counter += 1
            continue

        # copy matlab file to temp file
        shutil.copyfile(root_dir + '/' + inputFile, tempFile)

        # create/overwrite calibrations file
        with open(calibrationTxtFile, 'w+') as f4:
            f4.truncate(0)

        ### Data Manipulation ###

        # assign values to calibrations variable
        with open(tempFile, 'r') as f1:
            calibrations1 = f1.readlines()

        numLines = sum(1 for line in calibrations1) - 1

        # Strip Comments that came from Matlab File
        # Take any Cals that span more than 1 line and put onto 1 line
        # Write to CalibrationsTxtFile
        remove_comments(calibrations1, numLines)
        multi_line_cal_to_single_line(calibrations1, numLines)
        write_calibrations_file(calibrations1)

        # read in the file you just created
        with open(calibrationTxtFile, 'r+') as f3:
            cals_text = f3.readlines()

        # search for references
        cal_names_only, cal_array_by_file = search_cal_file_for_referenced_cal_names()
        arrays_of_all_file_cal_arrays.append(cal_array_by_file)

write_to_excel(arrays_of_all_file_cal_arrays, date_, time_)

print(f"\nAll Finished. Your calibrations are found in Calibration.xslx in the same folder as this .py file.\n")
if wrong_file_counter != 0:
    print(f"{wrong_file_counter} files in that folder were not .m files and were not read.")
os.system("start EXCEL.EXE Calibrations.xlsx")
