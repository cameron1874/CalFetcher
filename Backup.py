import shutil
# import numpy
import datetime
import os
# from tkinter import Tk
from tkinter.filedialog import askdirectory
# import xlsxwriter as xl
import openpyxl


def remove_values_from_list(the_list, val):
    return [value for value in the_list if value != val]


def removecomments(cals, numlines):
    for index in range(0, numlines):
        tmpstr = cals[index]
        stridx = tmpstr.find('%')
        if stridx != -1:
            tmpstr = tmpstr[0:stridx - 1] + '\n'
            cals[index] = tmpstr
    return cals


def setup_out_file():
    # Open Files in w+ mode-> reading and writing, existing content is cleared, a new empty
    # file is created if it doesn't exist, file pointer is positioned at the beginning
    with open(tempFile, 'w+') as file:
        pass

    with open(calibrationTxtFile, 'w+') as file:
        pass

    with open(outFile, 'w+') as f5:
        datestamp = datetime.datetime.now()
        date = datestamp.strftime("%m/%d/%Y")
        time = datestamp.strftime('%H:%M:%S\n')
        f5.write(date + time)
    return date, time

def multi_line_cal_to_single_line(calibrations, numlines):
    for index in range(numlines, 0, -1):
        curstr = calibrations[index]
        closedbracket = curstr.find(']')
        openbracket = curstr.find('[')
        count = 0
        # if closedBracket != -1 and openBracket == -1: # Detecting multiple lines (note reading from the bottom up)
        if closedbracket != -1 and openbracket == -1:  # Detecting multiple lines (note reading from the bottom up)
            while openbracket == -1:
                # move current index-count to index-count-1
                calibrations[index - count - 1] = calibrations[index - count - 1].strip() + ' ; ' + calibrations[
                    index - count]
                # remove data from index-count
                calibrations[index - count] = 'DELETE'

                # conditions for the next check
                count = count + 1
                curstr2 = calibrations[index - count]
                openbracket = curstr2.find('[')
            index = index - count
    calibrations = remove_values_from_list(calibrations, 'DELETE')


def write_calibrations_file(calibrations, numlines):
    numlines = sum(1 for line in calibrations) - 1
    with open(calibrationTxtFile, 'w') as f2:
        for index in range(0, numlines):
            f2.write(calibrations[index])


def search_cal_file_for_referenced_cal_names():

    cal_reference_array = []
    cal_array_per_file = []
    cal_names = []
    array = []

    with open(outFile, 'a') as f0:
        with open(referenceFile, 'r') as f2:
            f0.write('Data_based_on_input_file: ' + inputFile + '\n')
            for index in f2:
                cal_reference_array.append(index)
                cal_full_names = index.splitlines()
                for index2 in range(0, sum(1 for line in cals_text) - 1):
                    str2 = (cals_text[index2]).lower()
                    cal_text_indexes = str2.find((cal_full_names[0].strip()).lower())

                    # split header from string, header is separated by '.'
                    if '.' not in str2:
                        continue
                    else:
                        header, str2 = str2.split('.', 1)
                    split_str2 = str2.split()

                    if cal_text_indexes != -1:
                        # Make array containing all cleaned up cal names
                        if split_str2[0] in cal_names:
                            pass
                        else:
                            cal_names.append(split_str2[0])
                        f0.write(str2)
                        cal_array_per_file.append(str2)
                        flag = 1
                        break
                    else:
                        flag = 0
                if flag < 1:  # not found
                    cal_not_found_message_for_file = str(cal_full_names[0] + '=\tNOT_FOUND\n')
                    cal_not_found_message_for_list = str(cal_full_names[0] + '=NOT_FOUND')
                    cal_array_per_file.append(cal_not_found_message_for_list)
                    f0.write(cal_not_found_message_for_file)
                    array.append(cal_full_names[0])
                else:
                    flag = 0  # found string in nested loop
            f0.write('NEXT_FILE\n')
    return cal_reference_array, cal_names, cal_array_per_file


def size_columns_to_fit(worksheet):
    for ind, column in enumerate(worksheet.columns):
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        if ind == 0:
            adjusted_width = max_length  # * 1.2
        else:
            adjusted_width = 14
        worksheet.column_dimensions[column_letter].width = adjusted_width


def write_to_excel(cal_reference_array, array_of_all_file_cal_arrays, date, time):

    # Initializations vars needed for this function
    input_file = 'out.txt'
    first_row_count = 0
    workbook = openpyxl.Workbook()
    # worksheet1 = workbook.name('Calibrations',1)
    worksheet = workbook.worksheets[0]
    worksheet.title = 'Scalar Cals'
    array_indexes_all_files = []
    cal_arrays_inc_not_found = []
    array_of_scalar_cals = []
    scalar_cal_arrays_values_only = []
    scalar_cal_values_only = []
    scalar_cal_names = []
    split_array_rows = []
    array_cal_names = []
    calibration_arrays = []
    loop_counter = 0

    # Populate top of sheet with file names
    for i0, file in enumerate(files):
        worksheet.cell(3, 2 + i0, str(file))

    # Populate top of sheet with date and time
    worksheet.cell(1, 1, date)
    worksheet.cell(1, 2, time)

    # Loop through list of cal lists to find array cals, make list of indexes of array cals for each file
    for i1, cal_array_per_m_file in enumerate(array_of_all_file_cal_arrays):
        array_indexes_per_file = []
        for i2, calibration in enumerate(cal_array_per_m_file):
            #calibration = calibration.replace('=', '')
            #calibration = calibration.replace(';', '')
            if '[' and ']' in calibration:
                array_indexes_per_file.append(i2)
        array_indexes_all_files.append(array_indexes_per_file)

    # Find longest list of array cals (most will be the same as longest, this is to get
    # index values so calibrations that are NOT_FOUND at those same indexes can be included
    max_list = max(array_indexes_all_files, key=len)
    # Sort list in descending order so when we remove from those indexes, numbers don't shift
    max_list.sort(reverse=True)

    # Remove cals corresponding to cal arrays, using indexes collected above
    # Assemble list containing only scalar cals
    for cal_array_per_m_file in array_of_all_file_cal_arrays:
        temp_cal_array = cal_array_per_m_file
        # Enumerate backwards so index counts down and removes from high index first
        for i2, cal in reversed(list(enumerate(cal_array_per_m_file))):
            for i3, value in enumerate(max_list):
                if int(i2) == int(value):
                    cal_arrays_inc_not_found.append(cal)
                    temp_cal_array.pop(i2)
                    loop_counter += 1
                    # Lots of m files to go through but only need to remove indexes from cal names once
                    if loop_counter <= len(max_list):
                        cal_names.pop(i2)
        array_of_scalar_cals.append(temp_cal_array)

    # Split values for scalar cals off into their own list
    for scalar_cal_array in array_of_scalar_cals:
        scalar_temp_array_per_file = []
        for scalar_cal in scalar_cal_array:
            scalar_cal = scalar_cal.replace('\n', '')
            scalar_cal = scalar_cal.replace(';', '')
            split_scalar = scalar_cal.split('=')
            for item in split_scalar:
                item = item.replace('=', '')
                item.strip()
            scalar_temp_array_per_file.append(split_scalar[-1])
        scalar_cal_arrays_values_only.append(scalar_temp_array_per_file)

    # Populate Cal names into first column of first sheet
    for i4, cal_name in enumerate(cal_names):
        worksheet.cell(5+i4, 1, cal_name)

    # Populate scalar cal values into first sheet
    for i5, array_of_scalar_cal_values in enumerate(scalar_cal_arrays_values_only):
        for i6, scalar_cal_value in enumerate(array_of_scalar_cal_values):
            try:
                scalar_cal_value = round(float(scalar_cal_value),4)
            except:
                pass
            worksheet.cell(5+i6, 2+i5, scalar_cal_value)
    size_columns_to_fit(worksheet)








    """calibration = calibration.replace('[', '')
                calibration = calibration.replace(']', '')

                split_calibration_array = list(calibration.split())
                # Save name of cal that is an array to list array_cal_names if it is not in list already
                # then delete it from split_array_row var
                array_of_arrays_minus_array_cals = [array_of_all_file_cal_arrays]
                if split_calibration_array[0] in array_cal_names:
                    pass
                else:
                    array_cal_names.append(split_calibration_array[0])
                del split_calibration_array[0]
                split_array_rows.append(split_calibration_array)
            print(split_array_rows)"""




    """
    with open(input_file, 'r') as data:  # read in text mode
        for index, row in enumerate(data.readlines()):
            if '/' in row:
                worksheet.cell(1, 1, row)
                continue
            elif '.m' in row:
                continue
            elif len(row) < 2:
                row = [row]
                worksheet.append(row)
                continue
            row = row.replace('=', '')
            row = row.replace(';', '')
            # Write cals that are not arrays (no brackets) into first page of Excel
            if '[' and ']' in row:
                # Find Calibrations that are arrays
                row = row.replace('[', '')
                row = row.replace(']', '')
                split_array_row = list(row.split())
                # Save name of cal that is an array to list array_cal_names if it is not in list already
                # then delete it from split_array_row var
                if split_array_row[0] in array_cal_names:
                    pass
                else:
                    array_cal_names.append(split_array_row[0])
                del split_array_row[0]
                split_array_rows.append(split_array_row)

                # Operations to be performed if cal is not an array
            else:
                split_row = list(row.split())
                if first_row_count < (len(cal_reference_array) - len(array_cal_names)):
                    worksheet.append(split_row)
                    first_row_count += 1
                else:
                    for item in split_row:
                        if 'c' or 'k' in item:
                            split_row.remove(item)
                            #print(split_row)
                            # for i0, file in enumerate(files):
                            # worksheet.cell(2, 2+i0, file)
                            # upper_range = len(cal_reference_array) - len(array_cal_names)
                            # for i9 in range(0,upper_range):
                            # print(split_row)
                            # worksheet.cell(4+i9, 2, str(split_row))
                size_columns_to_fit(worksheet)"""

    """ # Iterate over Cal names list and create sheet for each one
        for i, cal in enumerate(array_cal_names):
            calibration_arrays.append(split_array_rows[i::len(array_cal_names)])

            # Excel sheets can only have 31char length Titles
            if len(cal) > 31:
                old_name = cal
                workbook.create_sheet(title=cal[:30], index=(i + 1))
                cal = old_name
            elif len(cal) <= 31:
                workbook.create_sheet(title=cal, index=(i + 1))

            worksheet_names = workbook.sheetnames
            worksheet = workbook[worksheet_names[i + 1]]
            cal = [cal]
            empty_line = [' ']
            worksheet.append(cal)
            worksheet.append(empty_line)

            for ind, filename in enumerate(files):
                filename = [filename]
                worksheet.append(filename)

            for i2, array in enumerate(calibration_arrays[i]):
                for i3, value in enumerate(array):
                    worksheet.cell(row=3 + i2, column=2 + i3, value=value)

            size_columns_to_fit(worksheet)
        else:
            worksheet.insert_rows(index)"""
    workbook.save('Calibrations.xlsx')


# Author: Cody Palmer
# Last modified by: Cameron Floyd
# Last modified Date: 3-14-2024

# There is a README.txt for instructions

rootdir = askdirectory(title='Select Folder that Contains your .m Files')

wrong_file_counter = 0

### File Preparation ###

### Initialize Files Needed ###

tempFile = 'temp.txt'
calibrationTxtFile = 'calibrations.txt'
outFile = 'out.txt'
referenceFile = 'references.txt'

array_of_all_file_cal_arrays = []

# create/overwrite output file and add timedate to the top
date, time = setup_out_file()

# loop to iterate through files in selected folder
for subdir, dirs, files in os.walk(rootdir):
    for inputFile in files:
        if str(inputFile).endswith('.py'):
            continue
        elif str(inputFile).endswith('.txt'):
            continue
        elif str(inputFile).endswith('.m') is not True:
            wrong_file_counter += 1
            continue

        # copy matlab file to temp file
        shutil.copyfile(rootdir + '/' + inputFile, tempFile)

        # create/overwrite calibrations file
        with open(calibrationTxtFile, 'w+') as f4:
            f4.truncate(0)

        ### Data Manipulation ###

        # assign values to calibrations variable
        with open(tempFile, 'r') as f1:
            calibrations = f1.readlines()

        numLines = sum(1 for line in calibrations) - 1

        # Strip Comments that came from Matlab File
        # Take any Cals that span more than 1 line and put onto 1 line
        # Write to CalibrationsTxtFile
        removecomments(calibrations, numLines)
        multi_line_cal_to_single_line(calibrations, numLines)
        write_calibrations_file(calibrations, numLines)

        # read in the file you just created
        with open(calibrationTxtFile, 'r+') as f3:
            cals_text = f3.readlines()

        # search for references
        cal_reference_array, cal_names, cal_array_per_file = search_cal_file_for_referenced_cal_names()
        array_of_all_file_cal_arrays.append(cal_array_per_file)



write_to_excel(cal_reference_array, array_of_all_file_cal_arrays, date, time)


print(f"\nAll Finished. Your calibrations are found in Calibration.xslx in the same folder as this .py file.\n"
      f"{wrong_file_counter} files in that folder were not .m files and were not read.")
